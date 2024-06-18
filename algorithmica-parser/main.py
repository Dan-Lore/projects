from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.hyperlink import Hyperlink

url = 'https://ru.algorithmica.org/'
response = requests.get(url)

if response.status_code == 200:
    html_content = response.content
    soup = BeautifulSoup(html_content, 'html.parser')

    wb = Workbook()
    ws = wb.active
    ws.append(['h3', 'section_name', 'li'])

    #......................................
    division_div = soup.find_all('div', class_='division')[0]

    if division_div:
        h3_sections_mapping = {}
        current_h3 = None

        for element in division_div.children:
            if element.name == 'h3':
                current_h3 = element
                h3_sections_mapping[current_h3] = []
            elif element.name == 'section' and current_h3:
                h3_sections_mapping[current_h3].append(element)

    #......................................

        for h3, sections in h3_sections_mapping.items():
            for section in sections:
                section_name = section.find('a').text
                list_items = section.find('ul').find_all('li')

                # Создадим массив строк с названиями и ссылками
                items_array = []
                for li in list_items:
                    item_name = li.find('a').text
                    item_link = li.find('a')['href']

                    cell = ws.cell(row=ws.max_row + 1, column=1, value=h3.text)
                    cell = ws.cell(row=ws.max_row, column=2, value=section_name)
                    cell = ws.cell(row=ws.max_row, column=3, value=item_name)
                    cell.font = Font(underline='single')
                    cell.hyperlink = Hyperlink(ref=item_link, target=item_link)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40 
    ws.column_dimensions['C'].width = 40

    excel_filename = 'parsed_data.xlsx'
    wb.save(excel_filename)
    print(f'Данные успешно записаны в файл {excel_filename}')
else:
    print('Страница недоступна')